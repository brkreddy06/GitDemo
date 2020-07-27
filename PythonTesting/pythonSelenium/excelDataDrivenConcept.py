import openpyxl  # first import this one
# load_workbook() will load the complete work book (excel sheet)
book = openpyxl.load_workbook("C:\\Users\\brkre\\Ram\\SeleniumPython\\DataDrivenExcel\\DataDrivenSeleniumPython.xlsx")
sheet = book.active  # active method considers TestData sheet in excel
# In Python each row treated starting from no.1 column also stats from 1
# cell method will have the control on the TestData sheet cells. Now it go to row1 & coloumn2
# To read the value from excel sheet and print on the console then line 9 & 10 will do
# value method will extract the data from TestData sheet
cell = sheet.cell(row=1, column=2)
print(cell.value)

#if you want to write the value back to the excel sheet then line 13 will do
sheet.cell(row=2, column=2).value = "Ramkumar"
print(sheet.cell(row=2, column=2).value)
#if you want to know how many rows data present in the excel use - sheet.max_row method
#if you want to know how many columns data present in the excel use - sheet.max_column method
print(sheet.max_row)
print(sheet.max_column)

#there is another way to print the excel cell value instead of using line no.9 & 10
print(sheet['A7'].value)  #A7 - row=8, column=1

#To print every value present in the excel sheet - TestData -> use range in FOR loop
for i in range(1, sheet.max_row+1):  #sheet.max_row+1 if u don't give +1 then it prints sheet.max_row-1 rows
    print(sheet.cell(row=i, column=1).value)  #the line prints only rows value present in column1

print("******************** print all the row and column values ******************")
#if you want to print first row and all the column values then use 2 FOR loops
for r in range(1, sheet.max_row+1):  #to get rows
    for c in range(1, sheet.max_column+1):  #to get columns
        print(sheet.cell(row=r, column=c).value)

print("*******************Only row 2 values ******************************")
# Requirement print/consider only row2 values in the excelsheet, use IF condition
for r1 in range(1, sheet.max_row+1):
    if sheet.cell(row=r1, column=1).value == "Testcase6":  #no need to check all the column values as we need TestCase1 values, so give column=1
        for c1 in range(2, sheet.max_column+1):  # if u give 2 in range it will skip printing Testcase6, if u want TC6 printed then give 1
            print(sheet.cell(row=r1, column=c1).value)


